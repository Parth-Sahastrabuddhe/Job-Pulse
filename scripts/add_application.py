"""Add a job application entry to the Excel tracker."""
import sys
import shutil
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

TRACKER_PATH = Path(r"C:/Users/sahas/OneDrive/Desktop/Resume/Full_time/Application - full time (new).xlsm")
BACKUP_PATH = TRACKER_PATH.with_suffix(".xlsm.bak")

STATUS_COLORS = {
    "Applied": "FFFFFF00",
    "Rejected": "FFFF0000",
    "Referral": "FFCCFFFF",
    "Interview": "FFADD8E6",
    "Assessment": "FFFFCC99",
    "Offer Letter": "FF90EE90",
}

def add_application(company, role, url):
    shutil.copy2(TRACKER_PATH, BACKUP_PATH)

    wb = load_workbook(TRACKER_PATH, keep_vba=True)
    ws = wb["Sheet1"]

    # Find next empty row (check column A)
    row = 1
    while ws.cell(row=row, column=1).value is not None:
        row += 1

    ws.cell(row=row, column=1, value=company)      # A: Company
    ws.cell(row=row, column=2, value=role)         # B: Role
    ws.cell(row=row, column=4, value=date.today().strftime("%#m/%d/%Y")) # D: Date Applied (e.g. 3/20/2026)
    ws.cell(row=row, column=5, value="Applied")    # E: Status
    ws.cell(row=row, column=7, value=url)          # G: URL

    # Apply formatting
    ws.cell(row=row, column=5).fill = PatternFill(patternType="solid", fgColor=STATUS_COLORS["Applied"])
    center = Alignment(horizontal="center", vertical="center")
    for col in range(1, 7):  # A-F centered
        ws.cell(row=row, column=col).alignment = center
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="left", vertical="center")

    wb.save(TRACKER_PATH)
    print(f"OK|{row}")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: add_application.py <company> <role> <url>", file=sys.stderr)
        sys.exit(1)
    add_application(sys.argv[1], sys.argv[2], sys.argv[3])
