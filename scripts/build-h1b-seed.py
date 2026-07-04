#!/usr/bin/env python3
"""Build the committed H-1B LCA seed JSON from DOL LCA disclosure xlsx files.

Runs LOCALLY (never on EC2: the quarterly disclosure files are ~100MB each and
parsing them needs more RAM/time than the box has to spare). Aggregates
certified H-1B LCAs per normalized employer name and emits a compact JSON that
src/h1b-sponsors-seed.js consumes (via the matching layer in
src/h1b-matching.js) to fill h1b_sponsors.lca_count / avg_salary.

Download the quarterly files from the DOL OFLC performance-data page
(LCA_Disclosure_Data_FY20XX_QY.xlsx), then:

  python3 scripts/build-h1b-seed.py --window 2025 --out seeds/h1b-lca.json \
      /path/lca_FY2025_Q2.xlsx /path/lca_FY2025_Q3.xlsx \
      /path/lca_FY2025_Q4.xlsx /path/lca_FY2026_Q1.xlsx

Method notes:
- Counts rows with CASE_STATUS starting "Certified" and VISA_CLASS == "H-1B"
  (excludes E-3 / H-1B1 rows present in the same files).
- Wages are annualized from WAGE_RATE_OF_PAY_FROM + WAGE_UNIT_OF_PAY; values
  outside $20k..$2M are excluded from the wage median but still counted.
- Employers with fewer than MIN_COUNT certified LCAs are dropped to keep the
  committed seed small; pass --full to also dump the unfiltered aggregate.
- The employer-name normalization here MUST stay in sync with
  normalizeEmployerName() in src/h1b-matching.js.
"""

import argparse
import json
import os
import re
import statistics
import sys
from datetime import datetime, timezone

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install --user openpyxl")

MIN_COUNT = 3

# Trailing legal-form tokens stripped iteratively (as whole words only).
LEGAL_SUFFIXES = {
    "INC", "LLC", "CORP", "CORPORATION", "CO", "COMPANY", "LTD", "LIMITED",
    "LP", "LLP", "PLC", "PC", "PLLC", "PBC", "INCORPORATED", "&",
}

UNIT_FACTORS = {
    "YEAR": 1.0,
    "HOUR": 2080.0,
    "WEEK": 52.0,
    "BI-WEEKLY": 26.0,
    "BIWEEKLY": 26.0,
    "MONTH": 12.0,
}

REQUIRED_COLUMNS = (
    "EMPLOYER_NAME",
    "CASE_STATUS",
    "VISA_CLASS",
    "WAGE_RATE_OF_PAY_FROM",
    "WAGE_UNIT_OF_PAY",
)


def normalize_employer(name):
    if not name:
        return ""
    s = str(name).upper()
    s = re.sub(r"[.,'\"()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Collapse spaced-out legal forms ("L L C", "L P") so the suffix strip
    # below catches them ("BLOOMBERG L P" -> "BLOOMBERG LP" -> "BLOOMBERG").
    s = re.sub(r"\bL L C\b", "LLC", s)
    s = re.sub(r"\bL L P\b", "LLP", s)
    s = re.sub(r"\bL P\b", "LP", s)
    parts = s.split(" ")
    while parts and parts[-1] in LEGAL_SUFFIXES:
        parts.pop()
    return " ".join(parts)


def parse_wage(raw, unit):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        value = float(raw)
    else:
        cleaned = re.sub(r"[$,\s]", "", str(raw))
        try:
            value = float(cleaned)
        except ValueError:
            return None
    factor = UNIT_FACTORS.get(str(unit or "").strip().upper())
    if not factor:
        return None
    annual = value * factor
    if annual < 20_000 or annual > 2_000_000:
        return None
    return annual


def scan_file(path, agg):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {}
    for want in REQUIRED_COLUMNS:
        try:
            idx[want] = header.index(want)
        except ValueError:
            sys.exit(f"{path}: missing column {want}; first headers: {header[:25]}")

    scanned = certified = 0
    for row in rows:
        scanned += 1
        if scanned % 100_000 == 0:
            print(f"  {os.path.basename(path)}: {scanned:,} rows...", file=sys.stderr, flush=True)
        status = str(row[idx["CASE_STATUS"]] or "")
        if not status.startswith("Certified"):
            continue
        if str(row[idx["VISA_CLASS"]] or "").strip() != "H-1B":
            continue
        name = normalize_employer(row[idx["EMPLOYER_NAME"]])
        if not name:
            continue
        certified += 1
        entry = agg.setdefault(name, [0, []])
        entry[0] += 1
        wage = parse_wage(row[idx["WAGE_RATE_OF_PAY_FROM"]], row[idx["WAGE_UNIT_OF_PAY"]])
        if wage is not None:
            entry[1].append(wage)
    wb.close()
    print(f"{os.path.basename(path)}: {scanned:,} scanned, {certified:,} certified H-1B", file=sys.stderr, flush=True)


def summarize(agg, min_count):
    out = {}
    for name, (count, wages) in agg.items():
        if count < min_count:
            continue
        out[name] = {"n": count, "w": int(statistics.median(wages)) if wages else 0}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--window", required=True, help='provenance label, e.g. "2025"')
    ap.add_argument("--full", default=None, help="optional path for the unfiltered aggregate")
    ap.add_argument("files", nargs="+")
    args = ap.parse_args()

    agg = {}
    for path in args.files:
        scan_file(path, agg)

    meta = {
        "window": args.window,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_files": [os.path.basename(p) for p in args.files],
        "min_count": MIN_COUNT,
        "employer_normalization": "upper, strip punctuation, strip trailing legal suffixes",
    }

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    filtered = summarize(agg, MIN_COUNT)
    with open(args.out, "w") as f:
        json.dump({"meta": meta, "employers": filtered}, f, separators=(",", ":"), sort_keys=True)
    print(f"wrote {args.out}: {len(filtered):,} employers (>= {MIN_COUNT} LCAs)", file=sys.stderr)

    if args.full:
        everything = summarize(agg, 1)
        with open(args.full, "w") as f:
            json.dump({"meta": meta, "employers": everything}, f, separators=(",", ":"), sort_keys=True)
        print(f"wrote {args.full}: {len(everything):,} employers (unfiltered)", file=sys.stderr)


if __name__ == "__main__":
    main()
